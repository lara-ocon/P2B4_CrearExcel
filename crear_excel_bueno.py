
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, BarChart3D, ProjectedPieChart, PieChart, LineChart
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def extract(fichero):
    return pd.read_csv(f'ficheros_csv/{fichero}')


def pie_chart(pestaña, min_col, min_row, max_row, titulo, posicion):
    # vamos a añadir un grafico de sectores con la media de las pizzas
    pie = PieChart()
    labels = Reference(pestaña, min_col=min_col[0], min_row=min_row[0], max_row=max_row[0])
    data = Reference(pestaña, min_col=min_col[1], min_row=min_row[1], max_row=max_row[1])
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = titulo

    pie.height = 20 # default is 7.5
    pie.width = 20

    # Cut the first slice out of the pie
    # slice = DataPoint(idx=0, explosion=20)
    # pie.series[0].data_points = [slice]

    pestaña.add_chart(pie, posicion)
    pie.style = 7


def bar_chart(pestaña, min_col, max_col, min_row, max_row, titulo, posicion, label_x, label_y, color):
    barchart = BarChart()

    # solo cogemos las categorias (la columna es la misma)
    categorias = Reference(pestaña, min_col=min_col[0], max_col=min_col[0], min_row=min_row[0], max_row=max_row[0])
    # le decimos que la data esta en esa pestaña, y le indicamos de donde a donde está
    data = Reference(pestaña, min_col=min_col[1], max_col=max_col[1], min_row=min_row[1], max_row=max_row[1])

    barchart.add_data(data, titles_from_data=True)

    barchart.set_categories(categorias)

    pestaña.add_chart(barchart, posicion) # le decimos que añada el barchart a la pestaña, y le decimos donde ponerlo
    barchart.title = titulo # titulo

    barchart.y_axis.title = label_y
    barchart.x_axis.title = label_x

    barchart.height = 10 # default is 7.5
    barchart.width = 40

    barchart.style = 5

    s = barchart.series[0]
    s.graphicalProperties.line.solidFill = "00000"
    s.graphicalProperties.solidFill = color



def line_chart(pestaña, min_col, max_col, min_row, max_row, titulo, posicion, label_x, label_y):
    # Chart with date axis
    c2 = LineChart()
    c2.title = titulo
    c2.style = 12
    c2.y_axis.title = label_y
    c2.y_axis.crossAx = 500
    c2.x_axis.title = label_x
    
    categorias = Reference(pestaña, min_col=min_col[0], max_col=min_col[0], min_row=min_row[0], max_row=max_row[0])
    # le decimos que la data esta en esa pestaña, y le indicamos de donde a donde está
    data = Reference(pestaña, min_col=min_col[1], max_col=max_col[1], min_row=min_row[1], max_row=max_row[1])


    c2.add_data(data, titles_from_data=True)
    c2.set_categories(categorias)

    pestaña.add_chart(c2, posicion)


def tabla_medias_pizzas(pestaña):
    # añadimos la formula media a la ultima fila de la tabla
    # vamos a añadir la media al final
    # para ello tendremos que llamar a cada celda por columna(letra) + fila(numero)
    # ord('A') = 65, ord('Z') = 90
    i = 65 # el valor que le sumamos para poner la 'A'
    letra_delante = ""
    letra = 66
    c = 0 # indica la columna

    pestaña[f"AJ5"] = "Pizza"
    pestaña[f"AK5"] = "Media"

    pestaña[f"AJ5"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    pestaña[f"AK5"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')

    while c < 32:
        # la fila donde lo vamos a poner es la 59
        columna = letra_delante+f"{chr(letra)}"
        # Calculamos el promedio de cada fila
        # print(f"{columna}59")
        pestaña[f"AJ{5+1+c}"] = pestaña[f"{columna}5"].value
        pestaña[f"AK{5+1+c}"] = f'=SUM({columna}5:{columna}58)/53'

        # resaltamos la celda con la media en naranja
        pestaña[f"AK{5+1+c}"].fill = PatternFill(start_color='ffa500', end_color='ffa500', fill_type='solid')

        #pestaña[f"{columna}59"].style = 5
        c += 1
        letra += 1
        if letra == 91:
            # cuando nos pasamos de la Z añadimos la A delante, hasta que nos volvamos a pasar and so on
            letra_delante = chr(i)
            i += 1
            letra = 65


def total_pizzas_semana(pestaña):
    pestaña['AH5'] = 'Total'
    c = 6
    while c < 59:
        pestaña[f'AH{c}'] = f'=SUM(B{c}:AG{c})'
        pestaña[f"AH{c}"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
        c += 1


# rojo = '00FF0000'
# naranja = 'ffa500'

def pestaña_pizzas_semanas(wb):

    pestaña = wb['Pizzas a la semana']

    pestaña['A1'] = "Reporte de pizzas pedidas a la semana"
    pestaña['A2'] = '2016'

    pestaña['A1'].font = Font(name = 'Arial', size = 20) # con color='00FF0000' se pone en rojo
    pestaña['A2'].font = Font(name = 'Arial', size = 14)


    tabla_medias_pizzas(pestaña) # añadimos una tabla con las medias de las pizzas pedidas por semana
    total_pizzas_semana(pestaña) # sumamos el total de pizzas que se piden cada semana
    pie_chart(pestaña, [36,37], [6,6], [37,37], 'Media de pizzas por tipo', "AM5")
    wb.save('reportev3.xlsx')
    # insertamos un grafico con como ha sido la evolucion de la pizza mas pedida = thai_ckn
    bar_chart(pestaña, [1, 31], [1, 31], [6, 5], [52, 52], "Evolución pedidos thai_ckn", 'B61', 'semanas', 'cantidad', '4169E1')

    # añadimos una ultima grafica que refleje la suma de pedidos por semana, para ello, cogemos la columna que sumaba los pedidos
    # por semana
    bar_chart(pestaña, [1, 34], [1, 34], [6, 5], [52, 52], "Evolución total pedidos", 'X61', 'semanas', 'numero pedidos', "ff9900")


def insertar_tabla_ingredientes(pestaña):
    # en esta funcion insertamos el dataframe con los ingredientes por semana en la pestaña que nos indiquen
    df_ingredientes_semana = extract('ingredientes_semana.csv').rename(columns={'Unnamed: 0': 'semana'}).set_index('semana')
    # vamos a insertar los ingredientes
    c = 0
    
    for r in dataframe_to_rows(df_ingredientes_semana, index=True, header=True):
        if c != 1: # no quiero q se iprima la palabra semana ahi
            pestaña.append(r)
        c+=1
    pestaña['A5'] = 'Semanas'
    
    # pintamos la celda con los ingredientes
    for rows in pestaña.iter_rows(min_row=5, max_row=5, min_col=2, max_col=66):
        for cell in rows:
            cell.fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    
    # pintamos tambien las semana
    for cols in pestaña.iter_cols(min_col=1, max_col=1, min_row=5, max_row=58):
        for cell in cols:
            cell.fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')


def añadir_cols_media_prediccion(pestaña):
    # con esta funcion añadiremos una tabla con la media de los ingredientes usados y la predicción
    # añadimos la formula media a la ultima fila de la tabla
    # vamos a añadir la media al final
    # para ello tendremos que llamar a cada celda por columna(letra) + fila(numero)
    # ord('A') = 65, ord('Z') = 90
    i = 65 # el valor que le sumamos para poner la 'A'
    letra_delante = ""
    letra = 66
    c = 0 # indica la columna

    pestaña[f"B62"] = "Ingrediente"
    pestaña[f"C62"] = "Media"
    pestaña[f"D62"] = "Predicción"

    pestaña[f"B62"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    pestaña[f"C62"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    pestaña[f"D62"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')

    """
    # recorremos el total de ingredientes
    while c < 65:
        # la fila donde lo vamos a poner es la 59
        columna = letra_delante+f"{chr(letra)}"
        print("columna", columna)

        # Calculamos el promedio de cada fila
        # print(f"{columna}59")
        pestaña[f"B{2+1+c}"] = pestaña[f"{columna}5"].value
        pestaña[f"C{2+1+c}"] = f'=SUM({columna}5:{columna}53)/65'
        pestaña[f"D{2+1+c}"] = f'=C{2+1+c}*1.2'

        # pintamos la celda con las medias de naranja claro
        pestaña[f"C{2+1+c}"].fill = PatternFill(start_color='ffa500', end_color='ffa500', fill_type='solid')
        # resaltamos la celda con la prediccion en naranja oscuro
        pestaña[f"D{2+1+c}"].fill = PatternFill(start_color='ff8c00', end_color='ff8c00', fill_type='solid')

        #pestaña[f"{columna}59"].style = 5
        c += 1
        letra += 1
        if letra == 91:
            # cuando nos pasamos de la Z añadimos la A delante, hasta que nos volvamos a pasar and so on
            letra_delante = chr(i)
            i += 1
            letra = 65
    """
    

def pestaña_ingredientes_semanas(wb):

    pestaña = wb['Ingredientes por semana']

    pestaña['A1'] = "Ingredientes necesitados por semana"
    pestaña['A2'] = '2016'

    pestaña['A1'].font = Font(name = 'Arial', size = 20) # con color='00FF0000' se pone en rojo
    pestaña['A2'].font = Font(name = 'Arial', size = 14)

    pestaña.append([]) # añadimos dos filas en blanco
    pestaña.append([])

    # insertamos la tabla con todos los ingredientes que se han necesitado a lo largo de las semanas
    insertar_tabla_ingredientes(pestaña)

    # ahora hacemos la media y la predicción



    # añadimos una tabla con la media de ingredientes necesitados y la predicción
    añadir_cols_media_prediccion(pestaña)


if __name__ == "__main__":

    # CREAMOS LA HOJA CON EL ANÁLISIS DE PIZZAS PEDIDAS

    # extraemos los dataframes que tenemos creados
    df_pizzas_semana = extract('pizzas_semana.csv').rename(columns={'Unnamed: 0': 'semana'}).set_index('semana')

    # creamos la tabla pivote y la insertamos en la hoja del excel: 'Pizzas a la semana'
    tabla_pivote = df_pizzas_semana.pivot_table(index='semana')
    tabla_pivote.to_excel('reportev4.xlsx', startrow=4, sheet_name='Pizzas a la semana')

    wb = load_workbook('reportev4.xlsx') # cargamos el workbook
    # writer = pd.ExcelWriter('reportev4.xlsx', engine = 'openpyxl')
    # writer.book = wb

    # añadimos una pagina con la informacion acerca de las pizzas que se han pedido a la semana
    pestaña_pizzas_semanas(wb)
    
    # guardamos por ahora
    wb.save('reportev4.xlsx')


    # AÑADIMOS OTRA HOJA CON LOS INGREDIENTES POR SEMANA ===============================================================#
    wb.create_sheet('Ingredientes por semana')

    pestaña_ingredientes_semanas(wb)

    # writer.save()
    # writer.close()

    wb.save('reportev5.xlsx')


