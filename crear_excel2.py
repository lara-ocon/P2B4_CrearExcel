
# Practica 2 Bloque 4 - Adquisición de datos - IMAT
# Reporte Ejcutivo Excel Pizzería Maven Pizzas
# Autor: Lara Ocón Madrid

"""
En esta práctica vamos a hacer uso de los dataframes que obtuvimos 
en las anteriores prácticas acerca de la pizzería Maven Pizzas, para
generar un reporte ejecutivo en Excel.
El Excel estará formado por 3 hojas:
- Hoja 1: Reporte ejecutivo.
- Hoja 2: Reporte de pedidos.
- Hoja 3: Reporte de ingredientes.
"""

# Importamos las librerías necesarias
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# creamos una función que genere la pagina 2 del reporte ejecutivo
# como ya hemos dicho esta contendrá un reporte de pedidos
def reporte_pedidos(wb):
    # creamos una ws
    ws = wb.create_sheet("Reporte de pedidos")

    # extraemos el dataframe que vamos a usar
    df_pizzas_semana = pd.read_csv('ficheros_csv/pizzas_semana.csv').rename(columns={'Unnamed: 0': 'semana'}).set_index('semana')

    # lo añadimos a una hoja del excel
    for row in df_pizzas_semana.iterrows():

        ws.append(list(row))
    






if __name__ == "__main__":

    # creamos el excel
    wb = openpyxl.Workbook()
    
    # creamos una función que genere la pagina 2 del reporte ejecutivo
    # como ya hemos dicho esta contendrá un reporte de pedidos
    reporte_pedidos(wb)
    
    # guardamos el excel
    wb.save('reporte_ejecutivo.xlsx')
