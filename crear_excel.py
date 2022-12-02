
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, PieChart, LineChart
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

def extract(fichero):
    return pd.read_csv(f'ficheros_csv/{fichero}')


def pie_chart(pestaña, min_col, min_row, max_row, titulo, posicion):
    # vamos a añadir un grafico de sectores con la media de las pizzas
    pie = PieChart()
    labels = Reference(pestaña, min_col=min_col[0], min_row=min_row[0], max_row=max_row[0])
    data = Reference(pestaña, min_col=min_col[1], min_row=min_row[1], max_row=max_row[1])
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = titulo

    pie.height = 20
    pie.width = 20

    pestaña.add_chart(pie, posicion)
    pie.style = 7


def bar_chart(pestaña, min_col, max_col, min_row, max_row, titulo, posicion, label_x, label_y, color):
    barchart = BarChart()

    # solo cogemos las categorias (la columna es la misma)
    categorias = Reference(pestaña, min_col=min_col[0], max_col=min_col[0], min_row=min_row[0], max_row=max_row[0])
    # le decimos que la data esta en esa pestaña, y le indicamos de donde a donde está
    data = Reference(pestaña, min_col=min_col[1], max_col=max_col[1], min_row=min_row[1], max_row=max_row[1])

    barchart.add_data(data, titles_from_data=True)

    barchart.set_categories(categorias)

    pestaña.add_chart(barchart, posicion) # le decimos que añada el barchart a la pestaña, y le decimos donde ponerlo
    barchart.title = titulo # titulo

    barchart.y_axis.title = label_y
    barchart.x_axis.title = label_x

    barchart.height = 10 # default is 7.5
    barchart.width = 40

    barchart.style = 5

    s = barchart.series[0]
    s.graphicalProperties.line.solidFill = "00000"
    s.graphicalProperties.solidFill = color



def line_chart(pestaña, min_col, max_col, min_row, max_row, titulo, posicion, label_x, label_y):
    # Chart with date axis
    c2 = LineChart()
    c2.title = titulo
    c2.style = 12
    c2.y_axis.title = label_y
    c2.y_axis.crossAx = 500
    c2.x_axis.title = label_x
    
    categorias = Reference(pestaña, min_col=min_col[0], max_col=min_col[0], min_row=min_row[0], max_row=max_row[0])
    # le decimos que la data esta en esa pestaña, y le indicamos de donde a donde está
    data = Reference(pestaña, min_col=min_col[1], max_col=max_col[1], min_row=min_row[1], max_row=max_row[1])


    c2.add_data(data, titles_from_data=True)
    c2.set_categories(categorias)

    pestaña.add_chart(c2, posicion)


def tabla_medias_pizzas(pestaña):
    # añadimos la formula media a la ultima fila de la tabla
    # vamos a añadir la media al final
    # para ello tendremos que llamar a cada celda por columna(letra) + fila(numero)
    # ord('A') = 65, ord('Z') = 90
    i = 65 # el valor que le sumamos para poner la 'A'
    letra_delante = ""
    letra = 66
    c = 0 # indica la columna

    pestaña[f"AJ5"] = "Pizza"
    pestaña[f"AK5"] = "Media"

    pestaña[f"AJ5"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    pestaña[f"AK5"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')

    while c < 32:
        # la fila donde lo vamos a poner es la 59
        columna = letra_delante+f"{chr(letra)}"

        # Calculamos el promedio de cada fila
        pestaña[f"AJ{5+1+c}"] = pestaña[f"{columna}5"].value
        pestaña[f"AK{5+1+c}"] = f'=SUM({columna}5:{columna}58)/53'

        # resaltamos la celda con la media en naranja
        pestaña[f"AK{5+1+c}"].fill = PatternFill(start_color='ffa500', end_color='ffa500', fill_type='solid')

        c += 1
        letra += 1
        if letra == 91:
            # cuando nos pasamos de la Z añadimos la A delante, hasta que nos volvamos a pasar and so on
            letra_delante = chr(i)
            i += 1
            letra = 65


def total_pizzas_semana(pestaña):
    pestaña['AH5'] = 'Total'
    c = 6
    while c < 59:
        pestaña[f'AH{c}'] = f'=SUM(B{c}:AG{c})'
        pestaña[f"AH{c}"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
        c += 1


def pestaña_pizzas_semanas(wb):

    pestaña = wb['Pizzas a la semana']

    pestaña['A1'] = "Reporte de pizzas pedidas a la semana"
    pestaña['A2'] = '2016'

    pestaña['A1'].font = Font(name = 'Arial', size = 20) # con color='00FF0000' se pone en rojo
    pestaña['A2'].font = Font(name = 'Arial', size = 14)

    tabla_medias_pizzas(pestaña) # añadimos una tabla con las medias de las pizzas pedidas por semana
    total_pizzas_semana(pestaña) # sumamos el total de pizzas que se piden cada semana
    pie_chart(pestaña, [36,37], [6,6], [37,37], 'Media de pizzas por tipo', "AM5")
   
    # insertamos un grafico con como ha sido la evolucion de la pizza mas pedida = thai_ckn
    bar_chart(pestaña, [1, 31], [1, 31], [6, 5], [52, 52], "Evolución pedidos thai_ckn", 'B61', 'semanas', 'cantidad', '4169E1')

    # añadimos una ultima grafica que refleje la suma de pedidos por semana, para ello, cogemos la columna que sumaba los pedidos
    # por semana
    bar_chart(pestaña, [1, 34], [1, 34], [6, 5], [52, 52], "Evolución total pedidos", 'X61', 'semanas', 'numero pedidos', "ff9900")


def insertar_tabla_ingredientes(pestaña):
    # en esta funcion insertamos el dataframe con los ingredientes por semana en la pestaña que nos indiquen
    df_ingredientes_semana = extract('ingredientes_semana.csv').rename(columns={'Unnamed: 0': 'semana'}).set_index('semana')
    # vamos a insertar los ingredientes
    c = 0
    
    for r in dataframe_to_rows(df_ingredientes_semana, index=True, header=True):
        if c != 1: # no quiero q se iprima la palabra semana ahi
            pestaña.append(r)
        c+=1
    pestaña['A5'] = 'Semanas'
    
    # pintamos la celda con los ingredientes
    for rows in pestaña.iter_rows(min_row=5, max_row=5, min_col=2, max_col=66):
        for cell in rows:
            cell.fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    
    # pintamos tambien las semana
    for cols in pestaña.iter_cols(min_col=1, max_col=1, min_row=5, max_row=58):
        for cell in cols:
            cell.fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')


def añadir_cols_media_prediccion(pestaña):
    # con esta funcion añadiremos una tabla con la media de los ingredientes usados y la predicción
    # añadimos la formula media a la ultima fila de la tabla
    # vamos a añadir la media al final
    # para ello tendremos que llamar a cada celda por columna(letra) + fila(numero)
    # ord('A') = 65, ord('Z') = 90
    i = 65 # el valor que le sumamos para poner la 'A'
    letra_delante = ""
    letra = 66
    c = 0 # indica la columna

    pestaña[f"B62"] = "Ingrediente"
    pestaña[f"C62"] = "Media"
    pestaña[f"D62"] = "Predicción"

    pestaña[f"B62"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    pestaña[f"C62"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    pestaña[f"D62"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')

    # recorremos el total de ingredientes
    while c < 65:
        # la fila donde lo vamos a poner es la 59
        columna = letra_delante+f"{chr(letra)}"

        # Calculamos el promedio de cada fila
        pestaña[f"B{62+1+c}"] = pestaña[f"{columna}5"].value
        pestaña[f"C{62+1+c}"] = f'=SUM({columna}5:{columna}58)/52'
        pestaña[f"D{62+1+c}"] = f'=C{62+1+c}*1.2'

        # pintamos la celda con las medias de naranja claro
        pestaña[f"C{62+1+c}"].fill = PatternFill(start_color='ffa500', end_color='ffa500', fill_type='solid')
        # resaltamos la celda con la prediccion en naranja oscuro
        pestaña[f"D{62+1+c}"].fill = PatternFill(start_color='ff8c00', end_color='ff8c00', fill_type='solid')

        c += 1
        letra += 1
        if letra == 91:
            # cuando nos pasamos de la Z añadimos la A delante, hasta que nos volvamos a pasar and so on
            letra_delante = chr(i)
            i += 1
            letra = 65


def barplot_prediccion(pestaña):
    # creamos un grafico de barras con la prediccion de ingredientes
    # esta se encunetra en la columna D, desde la fila 63 hasta la 127
    # la columna de los ingredientes esta en la columna B, desde la fila 63 hasta la 127

    # creamos la grafica
    grafica = BarChart()
    grafica.type = "col"
    grafica.style = 10
    grafica.title = "Predicción de ingredientes"
    grafica.y_axis.title = 'Cantidad'
    grafica.x_axis.title = 'Ingredientes'
    grafica.shape = 4

    # añadimos los datos
    data = Reference(pestaña, min_col=4, min_row=62, max_row=127)
    categories = Reference(pestaña, min_col=2, min_row=63, max_row=127)
    grafica.add_data(data, titles_from_data=True)
    grafica.set_categories(categories)
    grafica.height = 30
    grafica.width = 40

    # añadimos la grafica a la hoja
    pestaña.add_chart(grafica, "F63")


def lineplot_prediccion_vs_realidad(pestaña):
    # vamos a comparar la prediccion de Garlic con la realidad para cada semana
    # para ello vamos a crear un grafico de lineas
    # la realidad esta en la columna K, desde la fila 5 hasta la 58
    # la prediccion esta en la celda D72
    # queremos restarle a cada cantidad de Garlic de cada semana, la prediccion de Garlic
    # para ello vamos a crear una columna nueva, la columna AC, desde la fila 62 hasta la 127

    pestaña[f"AC62"] = "Semana"
    pestaña[f"AD62"] = "Realidad"
    pestaña[f"AE62"] = "Diferencia"

    pestaña[f"AC62"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    pestaña[f"AD62"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')
    pestaña[f"AE62"].fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')

    # rellenamos la columna AC con los numeros de semana
    i = 1
    while i < 54:
        pestaña[f"AC{62+i}"] = i                            # numero de semana
        pestaña[f"AD{62+i}"] = pestaña[f"K{5+i}"].value     # cantidad necesitada de Garlic esa semana
        pestaña[f"AE{62+i}"] = f'=D72-AD{62+i}'             # diferencia entre la prediccion y la realidad

        # resalatamos la celda con la cantidad en azul claro
        pestaña[f"AD{62+i}"].fill = PatternFill(start_color='add8e6', end_color='add8e6', fill_type='solid')
        # resaltamos la celda con la diferencia en azul oscuro
        pestaña[f"AE{62+i}"].fill = PatternFill(start_color='0000ff', end_color='0000ff', fill_type='solid')

        i += 1

    # creamos la grafica
    grafica = LineChart()
    grafica.type = "line"
    grafica.style = 10
    grafica.title = "Predicción vs Realidad Garlic"
    grafica.y_axis.title = 'Cantidad - Predicción'
    grafica.x_axis.title = 'Semana'
    grafica.shape = 4

    # añadimos los datos, estos son los de la tabla que acabamos de crear
    data = Reference(pestaña, min_col=31, min_row=62, max_row=62+53)
    categories = Reference(pestaña, min_col=29, min_row=63, max_row=62+53)
    grafica.add_data(data, titles_from_data=True)
    grafica.set_categories(categories)
    grafica.height = 30
    grafica.width = 40

    # añadimos la grafica a la hoja
    pestaña.add_chart(grafica, "AH63")



    

def pestaña_ingredientes_semanas(wb):

    pestaña = wb['Ingredientes por semana']

    pestaña['A1'] = "Ingredientes necesitados por semana"
    pestaña['A2'] = '2016'

    pestaña['A1'].font = Font(name = 'Arial', size = 20) # con color='00FF0000' se pone en rojo
    pestaña['A2'].font = Font(name = 'Arial', size = 14)

    pestaña.append([]) # añadimos dos filas en blanco
    pestaña.append([])

    # insertamos la tabla con todos los ingredientes que se han necesitado a lo largo de las semanas
    insertar_tabla_ingredientes(pestaña)

    # añadimos una tabla con la media de ingredientes necesitados y la predicción
    añadir_cols_media_prediccion(pestaña)

    # hacemos un barchart con la prediccion de los ingredientes
    barplot_prediccion(pestaña)

    # hacemos un lineplot con la prediccion vs la realidad de Garlic
    lineplot_prediccion_vs_realidad(pestaña)

    # finalmente añadimos un grafico de sectores con la media de los ingredientes
    pie_chart(pestaña, [2,3], [63,62], [127,127], "Cantidad de ingredientes necesitados", 'BE63')

def pestaña_reporte_ejecutivo(wb):
    # esta es la ultima pestaña del reporte ejecutivo
    # en esta pestaña vamos a hacer un resumen de los datos que hemos obtenido

    pestaña = wb['Reporte ejecutivo']

    pestaña['A1'] = "Ganancias"
    pestaña['A2'] = '2016'

    pestaña['A1'].font = Font(name = 'Arial', size = 20) # con color='00FF0000' se pone en rojo
    pestaña['A2'].font = Font(name = 'Arial', size = 14)

    pestaña.append([]) # añadimos dos filas en blanco
    pestaña.append([])

    # Insertamos la imagen con las ganancias al mes y alas ganancias al año
    img = Image('imagenes/ganancias_semana.png')
    pestaña.add_image(img, 'B5')

    img2 = Image('imagenes/ganancias_mes.png')
    pestaña.add_image(img2, 'Q5')


if __name__ == "__main__":

    # CREAMOS LA HOJA CON EL ANÁLISIS DE PIZZAS PEDIDAS =======================================================#

    # extraemos los dataframes que tenemos creados
    df_pizzas_semana = extract('pizzas_semana.csv').rename(columns={'Unnamed: 0': 'semana'}).set_index('semana')

    # creamos la tabla pivote y la insertamos en la hoja del excel: 'Pizzas a la semana'
    tabla_pivote = df_pizzas_semana.pivot_table(index='semana')
    tabla_pivote.to_excel('reporte_ejecutivo.xlsx', startrow=4, sheet_name='Pizzas a la semana')

    wb = load_workbook('reporte_ejecutivo.xlsx') # cargamos el workbook
    # writer = pd.ExcelWriter('reportev4.xlsx', engine = 'openpyxl')
    # writer.book = wb

    # añadimos una pagina con la informacion acerca de las pizzas que se han pedido a la semana
    pestaña_pizzas_semanas(wb)

    # AÑADIMOS OTRA HOJA CON LOS INGREDIENTES POR SEMANA =======================================================#
    wb.create_sheet('Ingredientes por semana')

    pestaña_ingredientes_semanas(wb)

    # AÑADIMOS UNA ULTIMA HOJA CON EL REPORTE EJECUTIVO =========================================================#
    wb.create_sheet('Reporte ejecutivo')

    pestaña_reporte_ejecutivo(wb)

    # Hemos terminado de crear el reporte ejecutivo, ahora lo guardamos
    wb.save('reporte_ejecutivo.xlsx')


